from openpyxl import load_workbook

# Load the workbook
file_path = "status1.xlsx"
wb = load_workbook(file_path)

# Select the active worksheet (or specify a sheet name)
ws = wb.active  # or ws = wb["SheetName"]

# Find the last row with data
last_row = ws.max_row  # Returns the last row that contains data

# Insert a new row after the last row
new_row_data = ["S no.","pcbid","datetime","info","imei","size","rtc","ptsc","pv","nw","wdt","bv","bvv","sd","bme","bmev","hm","remark","accept/reject"]  # Add your data here
ws.append(new_row_data)

# Save the workbook
wb.save(file_path)

print(f"New row added after row {last_row}!")
