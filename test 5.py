import tkinter as tk
from tkinter import StringVar, ttk,filedialog
import serial
import serial.tools.list_ports
import threading
from datetime import datetime
import logging
import subprocess
import sys
from tkinter import ttk
import uuid
from openpyxl import Workbook, load_workbook
import os


def updater():
    global values
    pmin=pressure_min_entry.get()
    pmax=pressure_max_entry.get()
    tmin=temp_min_entry.get()
    tmax=temp_max_entry.get()
    bpmin=bme_pressure_min_entry.get()
    bpmax=bme_pressure_max_entry.get()
    btmin=bme_temp_min_entry.get()
    btmax=bme_temp_max_entry.get()
    bamin=bme_altitude_min_entry.get()
    bamax=bme_altitude_max_entry.get()
    vmin=battery_voltage_min_entry.get()
    vmax=battery_voltage_max_entry.get()
    print(values)

    if(values["wdt"]=="PASS"):
        wdt_accept()
        values["finalwdt"]="PASS"
    else:
        wdt_reject()
        values["finalwdt"]="FAIL"

    if(values["info"]!="--"):
        info_accept()
        values["espimei"]=values["info"][0]
        values["espsize"]=values["info"][1]
        values["espmac"]=values["info"][2]
        values["finalinfo"]="PASS"
    else:
        info_reject
        values["finalinfo"]="FAIL"
        values["espimei"]=""
        values["espsize"]=""
        values["espmac"]=""

    if(values["rtc"]=="PASS"):
        rtc_accept()
        values["finalrtc"]="PASS"
    else:
        rtc_reject()
        values["finalrtc"]="FAIL"


    if(values["pressure"]=="FAIL"): 
        pressure_reject()
        values["ppressure"]=""
        values["ptemp"]=""
        values["finalpressure"]="FAIL"
    else:
        if(values["pressure"]!="--"):
            pressure_accept()
            values["ppressure"]=values["pressure"][0]
            values["ptemp"]=values["pressure"][1]
            values["finalpressure"]="PASS"
        else:
            pressure_reject()
            values["ppressure"]=""
            values["ptemp"]=""
            values["finalpressure"]="FAIL"

    if(values["battery"]=="FAIL"): 
        battery_reject()
        values["finalbattery"]="FAIL"
        values["voltage"]=""
    else:
        if(values["battery"]!="--"):
            battery_accept()
            values["finalbattery"]="PASS"
            values["voltage"]=values["battery"][0]
        else:
            battery_reject()
            values["finalbattery"]="FAIL"
            values["voltage"]=""

    if(values["sd"]=="PASS"):
        sd_accept()
        values["finalsd"]="PASS"
    else:
        sd_reject
        values["finalsd"]="FAIL"

    if(values["bme"]=="FAIL"): 
        pressure_reject()
        values["bpressure"]=""
        values["btemp"]=""
        values["baltitude"]=""
        values["finalbme"]="FAIL"
    else:
        if(values["bme"]!="--"):
            bme_accept()
            values["bpressure"]=values["bme"][0]
            values["btemp"]=values["bme"][1]
            values["baltitude"]=values["bme"][2]
            values["finalbme"]="PASS"
        else:
            bme_reject()
            values["bpressure"]=""
            values["btemp"]=""
            values["baltitude"]=""
            values["finalbme"]="FAIL"

    if(values["net"]=="PASS"):
        net_accept()
        values["finalnet"]="PASS"
    else:
        net_reject()
        values["finalnet"]="FAIL"

    if(values["hibernate"]=="PASS"):
        hibernate_accept()
        values["finalhib"]="PASS"
    else:
        hibernate_reject()
        values["finalhib"]="FAIL"

    print(values)

def wdt_accept():
    text10.config(bg="lightgreen")
    text11.config(bg="lightgreen")
    text12.config(bg="lightgreen",text="Accepted")
    
def wdt_reject():
    text10.config(bg="lightcoral")
    text11.config(bg="lightcoral")
    text12.config(bg="lightcoral",text="Rejected")

def info_accept():
    text20.config(bg="lightgreen")
    text21.config(bg="lightgreen")
    text22.config(bg="lightgreen",text="Accepted")

def info_reject():
    text20.config(bg="lightcoral")
    text21.config(bg="lightcoral")
    text22.config(bg="lightcoral",text="Rejected")

def rtc_accept():
    text30.config(bg="lightgreen")
    text31.config(bg="lightgreen")
    text32.config(bg="lightgreen",text="Accepted")

def rtc_reject():
    text30.config(bg="lightcoral")
    text31.config(bg="lightcoral")
    text32.config(bg="lightcoral",text="Rejected")

def pressure_accept():
    text40.config(bg="lightgreen")
    text41.config(bg="lightgreen")
    text42.config(bg="lightgreen",text="Accepted")

def pressure_reject():
    text40.config(bg="lightcoral")
    text41.config(bg="lightcoral")
    text42.config(bg="lightcoral",text="Rejected")

def battery_accept():
    text50.config(bg="lightgreen")
    text51.config(bg="lightgreen")
    text52.config(bg="lightgreen",text="Accepted")

def battery_reject():
    text50.config(bg="lightcoral")
    text51.config(bg="lightcoral")
    text52.config(bg="lightcoral",text="Rejected")

def sd_accept():
    text60.config(bg="lightgreen")
    text61.config(bg="lightgreen")
    text62.config(bg="lightgreen",text="Accepted")

def sd_reject():
    text60.config(bg="lightcoral")
    text61.config(bg="lightcoral")
    text62.config(bg="lightcoral",text="Rejected")

def bme_accept():
    text70.config(bg="lightgreen")
    text71.config(bg="lightgreen")
    text72.config(bg="lightgreen",text="Accepted")

def bme_reject():
    text70.config(bg="lightcoral")
    text71.config(bg="lightcoral")
    text72.config(bg="lightcoral",text="Rejected")

def net_accept():
    text80.config(bg="lightgreen")
    text81.config(bg="lightgreen")
    text82.config(bg="lightgreen",text="Accepted")

def net_reject():
    text80.config(bg="lightcoral")
    text81.config(bg="lightcoral")
    text82.config(bg="lightcoral",text="Rejected")

def hibernate_accept():
    text90.config(bg="lightgreen")
    text91.config(bg="lightgreen")
    text92.config(bg="lightgreen",text="Accepted")

def hibernate_reject():
    text90.config(bg="lightcoral")
    text91.config(bg="lightcoral")
    text92.config(bg="lightcoral",text="Rejected")

# Function to update the time
def update_timer():
    global trunning, counter
    while trunning:
        minutes, seconds = divmod(counter, 60)
        timer_text = f"Timer: {minutes:02}:{seconds:02}"
        time_label.config(text=timer_text)
        app.update()
        counter += 1
        threading.Event().wait(1)

def get_mac_address():
    mac = uuid.getnode()
    mac_address = ':'.join(f'{(mac >> i) & 0xff:02x}' for i in range(0, 8 * 6, 8)[::-1])
    return mac_address

#{'wdt': 'PASS', 'info': '862749051208634,8388608,558993728', 
# 'rtc': 'PASS', 'pressure': '1.012,18.020', 'battery': '3.052', 
# 'sd': 'PASS', 'bme': '19.760,0.985,342.433', 'net': 'FAIL', 
# 'hibernate': 'PASS', 'pcbid': 'sdjkbf', 'sourceid': 'd2:8c:af:f3:c4:43', 
# 'date': '2024-12-16', 'time': '13:41:31', 'finalwdt': 'PASS', 
# 'espimei': '8', 'espsize': '6', 'espmac': '2', 'finalinfo': 'PASS', 
# 'finalrtc': 'PASS', 'ppressure': '1', 'ptemp': '.', 'finalpressure': 'PASS',
#  'finalbattery': 'PASS', 'voltage': '3.052', 'finalsd': 'PASS', 
# 'bpressure': '', 'btemp': '', 'baltitude': '', 'finalbme': 'FAIL', 
# 'finalnet': 'FAIL', 'finalhib': 'PASS'}

def passtest():
    label5.config(state=tk.DISABLED)
    label4.config(fg="green")
    file_path = "output.xlsx"
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Date", "Time", "Pcb Id","Source Id","ESP32 IMEI","ESP32 SIZE","ESP32 Info Communication","Watch Dog Timer","RTC Communication","Pressure","Temperature","Pressure Sensor Communication","Battery Voltage","Battery Communication","SD Card Communication","Pressure","Temperature","Altitude","BME Sensor Communcation","GPRS Communication","Hibernate Mode Testing","Comments","Final Status"])
    row_to_insert = [values.get("date"), values.get("time"), values.get("pcbid"),values.get("sourceid"),values.get("espimei"),values.get("espsize"),values.get("finalinfo"),values.get("finalwdt"),values.get("finalrtc"),values.get("ppressure"),values.get("ptemp"),values.get("finalpressure"),values.get("voltage"),values.get("finalbattery"),values.get("finalsd"),values.get("bpressure"),values.get("btemp"),values.get("baltitude"),values.get("finalbme"),values.get("finalnet"),values.get("finalhib"),remark_var.get(),"PASS"]
    sheet.append(row_to_insert)
    wb.save(file_path)
    output_area.insert(tk.END, f"Data successfully written to {file_path}.\n")
    print(f"Data successfully written to {file_path}.")


def failtest():
    label4.config(state=tk.DISABLED)
    label5.config(fg="red")
    file_path = "output.xlsx"
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Date", "Time", "Pcb Id","Source Id","ESP32 IMEI","ESP32 SIZE","ESP32 Info Communication","Watch Dog Timer","RTC Communication","Pressure","Temperature","Pressure Sensor Communication","Battery Voltage","Battery Communication","SD Card Communication","Pressure","Temperature","Altitude","BME Sensor Communcation","GPRS Communication","Hibernate Mode Testing","Comments","Final Status"])
    row_to_insert = [values.get("date"), values.get("time"), values.get("pcbid"),values.get("sourceid"),values.get("espimei"),values.get("espsize"),values.get("finalinfo"),values.get("finalwdt"),values.get("finalrtc"),values.get("ppressure"),values.get("ptemp"),values.get("finalpressure"),values.get("voltage"),values.get("finalbattery"),values.get("finalsd"),values.get("bpressure"),values.get("btemp"),values.get("baltitude"),values.get("finalbme"),values.get("finalnet"),values.get("finalhib"),remark_var.get(),"FAIL"]
    sheet.append(row_to_insert)
    wb.save(file_path)
    output_area.insert(tk.END, f"Data successfully written to {file_path}.\n")
    print(f"Data successfully written to {file_path}.")


def setup_logging(pcbid):
    global txt_filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    txt_filename = f"Finalstatus_{pcbid}_{timestamp}.txt"

    logging.basicConfig(
        filename=txt_filename,
        level=logging.DEBUG,  # You can change the level to INFO or ERROR depending on what you need
        format="%(asctime)s - %(message)s",  # You can adjust the format as needed
    )

def write_flash():
    if not is_port_running:
        output_area.insert(tk.END, f"No port connected.\n")
        return

    file_path = file_var.get()
    if not (file_path):
        output_area.insert(tk.END, f"Please provide a file.\n")
        return

    command = [
        "python3", get_esptool_path(), '-p', port_var.get(), '-b', '460800', 
        '--before', 'default_reset', '--after', 'hard_reset', '--chip', 'esp32',
        'write_flash', '--flash_mode', 'dio', '--flash_size', 'detect', '--flash_freq', '40m',
        "0x0000", file_path
    ]
    run_in_thread(command)

# Function to run a command in a thread
def run_in_thread(command, on_complete=None):
    def target():
        try:
            output_area.insert(tk.END,f"Running command: {' '.join(command)}")
            process = subprocess.Popen(
                command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True
            )
            output_lines = []
            for line in iter(process.stdout.readline, ""):
                output_lines.append(line.strip())
            _, stderr = process.communicate()
            if stderr:
                output_area.insert(tk.END,f"Error: {stderr.strip()}")
            if on_complete:
                on_complete(output_lines)
        except Exception as e:
            output_area.insert(tk.END,f"Exception occurred: {e}")

    threading.Thread(target=target).start()

# Function to erase flash memory
def erase_flash():
    if not is_port_running:
        output_area.insert(tk.END, f"No port connected.\n")
        return
    command = ["python3", get_esptool_path(), '--chip', 'esp32', '--port', port_var.get(), 
               'erase_flash', '--force']
    run_in_thread(command)

# Function to read flash info
def read_flash_info():
    if not is_port_running:
        output_area.insert(tk.END, f"No port connected.\n")
        return
    command = ["python3", get_esptool_path(), '--port', port_var.get(), 'flash_id']
    run_in_thread(command)


def browse_file(file_var):
    file_path = filedialog.askopenfilename()
    if file_path:
        file_var.set(file_path)
        output_area.insert(tk.END, f"Selected file: {file_path}")

def enable_start_button(*args):
    if (pcbid_var.get() and is_port_running ==True):
        start_button.config(state=tk.NORMAL)
        stop_button.config(state=tk.DISABLED)

# Function to get the path of esptool.py
def get_esptool_path():
    # Check if running as a packaged executable with PyInstaller
    if getattr(sys, 'frozen', False):
        # If bundled with PyInstaller, use sys._MEIPASS for the temporary folder
        esptool_directory = os.path.join(sys._MEIPASS, "esptool-master")
    else:
        # If running as a normal script, use the directory of the script
        esptool_directory = os.path.join(os.path.dirname(__file__), "esptool-master")
    
    if not os.path.exists(esptool_directory):
        output_area.insert(tk.END,"esptool directory does not exist")
        raise FileNotFoundError(f"esptool directory does not exist: {esptool_directory}")
    
    return os.path.join(esptool_directory, "esptool.py")

# Function to get available COM ports
def get_ports():
    return [port.device for port in serial.tools.list_ports.comports()]

def start_timer():
    global timer_flag
    timer_flag = False  # Reset the flag at the start of the timer

    def timer_task():
        global timer_flag
        # Wait for 1.5 minutes (90 seconds)
        threading.Event().wait(90)
        print("hogyaa hu bhai")
        timer_flag = True  # Set the flag after the timer expires

    # Start the timer in a separate thread
    threading.Thread(target=timer_task, daemon=True).start()
    print("Timer started for 1.5 minutes.")

# Function to refresh the port dropdown
def refresh_ports():
    port_combobox['values'] = get_ports()
    if port_combobox['values']:
        port_combobox.current(0)

def connect():
    global ser, is_port_running
    port = port_var.get()
    
    connect_button.config(state=tk.DISABLED)
    refresh_button.config(state=tk.DISABLED)

    if not port:
        output_area.insert(tk.END, "Please select a COM port.\n")
        return
    
    output_area.insert(tk.END, "Connecting...\n")

    try:
        ser = serial.Serial(port, 115200, timeout=1)
        ser.flushInput()
        is_port_running = True
        output_area.insert(tk.END, f"Connected to {port} at 115200 baud.\n")
        threading.Thread(target=read_from_serial, daemon=True).start()
    except Exception as e:
        output_area.insert(tk.END, f"Failed to connect: {e}\n")
        is_port_running = False
        connect_button.config(state=tk.NORMAL)

def startover():
    global is_port_running,ser,datee,timee,values,hflag,switchflag
    global trunning, counter
    trunning = False
    counter = 0
    time_label.config(text="Timer: 00:00")
    ts.config(text="Test Status: --",font=("Poppins", 12, "bold"))
    if txt_filename:
        txt_filename.close()
        output_area.insert(tk.END, "Logging interrupted and saved\n")
    connect_button.config(state=tk.NORMAL)
    refresh_button.config(state=tk.NORMAL)
    label4.config(state=tk.NORMAL)
    label5.config(state=tk.NORMAL)
    label4.config(fg="black")
    label5.config(fg="black")
    is_port_running = False
    output_area.insert(tk.END, f"Port Disconnected!!\n")
    ser = None
    datee=None
    hflag=False
    mac=None
    values={"wdt":"--","info":"--","rtc":"--","pressure":"--","battery":"--","sd":"--","bme":"--","net":"--","hibernate":"--"}
    timee=None
    start_button.config(state=tk.DISABLED)
    stop_button.config(state=tk.DISABLED)
    open_window_button.config(state=tk.NORMAL)
    pcbid_var.set("")
    switchflag=True
    remark_var.set("")
    label1.config(bg="white")
    label2.config(bg="white")
    text10.config(bg="lightgrey")
    text20.config(bg="lightgrey")
    text30.config(bg="lightgrey")
    text40.config(bg="lightgrey")
    text50.config(bg="lightgrey")
    text60.config(bg="lightgrey")
    text70.config(bg="lightgrey")
    text80.config(bg="lightgrey")
    text90.config(bg="lightgrey")
    text11.config(text="--",bg="lightgrey")
    text21.config(text="--",bg="lightgrey")
    text31.config(text="--",bg="lightgrey")
    text41.config(text="--",bg="lightgrey")
    text51.config(text="--",bg="lightgrey")
    text61.config(text="--",bg="lightgrey")
    text71.config(text="--",bg="lightgrey")
    text81.config(text="--",bg="lightgrey")
    text91.config(text="--",bg="lightgrey")
    text12.config(text="--",bg="lightgrey")
    text22.config(text="--",bg="lightgrey")
    text32.config(text="--",bg="lightgrey")
    text42.config(text="--",bg="lightgrey")
    text52.config(text="--",bg="lightgrey")
    text62.config(text="--",bg="lightgrey")
    text72.config(text="--",bg="lightgrey")
    text82.config(text="--",bg="lightgrey")
    text92.config(text="--",bg="lightgrey")


def read_from_serial():
    global ser, is_port_running,wdtflag,qcflag,hflag,values,switchflag
    while is_port_running:
        try:
            if ser.in_waiting > 0:
                data = ser.readline().decode('utf-8', errors='ignore').strip()
                output_area.insert(tk.END, f"{data}\n")
                output_area.see(tk.END)  # Auto-scroll to the bottom
                
                if data == "[ExC Main] PSUP_ON"and switchflag==True:
                    output_area.insert(tk.END, f"Switch to PCB Supply, remove USB-C and connect the debug cable\n")
                    switchflag=False

                if data == "[ExC: Main] ESP_WDT_START_ACK":  
                    output_area.insert(tk.END, "Received ESP_WDT_START_ACK.\n")

                if data.startswith("[ExC: Main] WDT:"):
                    logging.info("WDT Test Starts...")
                    wdtflag=True

                if data.startswith("[ExC Main] QC Firmware") and wdtflag==True:
                    logging.info("[ExC: Main] WDT: Pass")
                    values["wdt"]="PASS"
                    text11.config(text="PASS",bg="lightgrey")
                    wdtflag=False
                    ser.write("ESP_QC_START\n".encode('utf-8'))  

                if wdtflag==True and data.startswith("[ExC: Main] WDT: Failed"):
                    logging.info("[ExC: Main] WDT: Fail")
                    text11.config(text="Fail",bg="lightgrey")
                    values["wdt"]="FAIL"
                    wdtflag=False
                    ser.write("ESP_QC_START\n".encode('utf-8')) 

                if data.startswith("[ExC: Main] ESP_START_ACK"):
                    qcflag=True
                    pfail=False
                    bfail=False
                    logging.info("QC Tests Starts...")

                if data.startswith("[ExC: Main] ESPinfo:") and qcflag==True:
                    oldvals=data[21:].strip()
                    print(type(oldvals))
                    vals= oldvals.split(',')
                    print(type(vals))
                    values["info"]=vals
                    text21.config(text=f"{oldvals}",bg="lightgrey")
                    logging.info(f"{data}")

                if data.startswith("[ExC: Main] RTCInterface: Pass") and qcflag==True:
                    text31.config(text="Pass",bg="lightgrey")
                    logging.info(f"{data}")
                    values["rtc"]="PASS"

                if data.startswith("[ExC: Main] RTCInterface: Failed") and qcflag==True:
                    text31.config(text="Fail",bg="lightgrey")
                    logging.info(f"{data}")
                    values["rtc"]="FAIL"

                if data=="[ExC: Main] PressureTransmitter: Failed" and qcflag==True:
                    text41.config(text="Fail",bg="lightgrey")
                    logging.info(f"{data}")
                    values["pressure"]="FAIL"
                    pfail=True

                if data.startswith("[ExC: Main] PressureTransmitter:") and qcflag==True and pfail==False:
                    oldvals=data[33:].strip()
                    print(type(oldvals))
                    vals= oldvals.split(',')
                    print(type(vals))
                    text41.config(text=f"{oldvals}",bg="lightgrey")
                    logging.info(f"{data}")
                    values["pressure"]=vals
                    pfail=True

                if data.startswith("[ExC: Main] BatteryVoltage:") and qcflag==True:
                    oldvals=data[28:].strip()
                    print(type(oldvals))
                    vals= oldvals.split(',')
                    print(type(vals))
                    values["battery"]=vals
                    text51.config(text=f"{oldvals}",bg="lightgrey")
                    logging.info(f"{data}")

                if data == "[ExC: Main] uSD: Pass" and qcflag==True:
                    text61.config(text="Pass",bg="lightgrey")
                    values["sd"]="PASS"
                    logging.info(f"{data}")

                if data == "[ExC: Main] uSD: Failed" and qcflag==True:
                    text61.config(text="Fail",bg="lightgrey")
                    values["sd"]="FAIL"
                    logging.info(f"{data}")

                if data=="[ExC: Main] BME: Failed" and qcflag==True:
                    text71.config(text="Fail",bg="lightgrey")
                    logging.info(f"{data}")
                    values["bme"]="FAIL"
                    bfail=True

                if data.startswith("[ExC: Main] BME:") and qcflag==True and bfail==False:
                    oldvals=data[17:].strip()
                    print(type(oldvals))
                    vals= oldvals.split(',')
                    print(type(vals))
                    values["bme"]=vals
                    text71.config(text=f"{oldvals}",bg="lightgrey")
                    logging.info(f"{data}")
                    bfail=True

                if data == "[ExC: Main] GPRSConnection: Pass" and qcflag==True:
                    text81.config(text="Pass",bg="lightgrey")
                    values["net"]="PASS"
                    logging.info(f"{data}")
                    qcflag=False
                    hflag=True
                    ser.write("ESP_LPM_START\n".encode('utf-8')) 

                if data == "[ExC: Main] GPRSConnection: Failed" and qcflag==True:
                    text81.config(text="Fail",bg="lightgrey")
                    values["net"]="FAIL"
                    logging.info(f"{data}")
                    qcflag=False
                    hflag=True
                    ser.write("ESP_LPM_START\n".encode('utf-8')) 

                if data=="[ExC: Main] ESP_LPM_ACK" and hflag==True:
                    output_area.insert(tk.END, f"lpm acknowledges\n")
                
                if data=="[ExC Main] QC Firmware" and hflag==True:
                    text91.config(text="Pass",bg="lightgrey")
                    logging.info(f"LPM Pass")
                    values["hibernate"]="PASS"
                    hflag=False
                    stop_test()

                if timer_flag==True:
                    print("aara to h")
                    values["hibernate"]="FAIL"
                    text91.config(text="Fail",bg="lightgrey")
                    logging.info(f"LPM Pass")
                    hflag=False
                    stop_test()
                
        except Exception as e:
            output_area.insert(tk.END, f"Error reading from serial: {e}\n")
            output_area.see(tk.END)
            break

def start_test():
    global ser,datee,timee,txt_filename,mac
    global trunning
    if not trunning:
        trunning = True
        threading.Thread(target=update_timer, daemon=True).start()
    label1.config(bg="lightgreen")
    label2.config(bg="white")
    start_button.config(state=tk.DISABLED)
    stop_button.config(state=tk.NORMAL)
    open_window_button.config(state=tk.DISABLED)
    if not ser or not ser.is_open:
        output_area.insert(tk.END, "No active serial connection. Startover and Connect to a port first.\n")
        return
    pcb_id = pcbid_var.get()
    mac= get_mac_address()
    values["pcbid"]=pcb_id
    values["sourceid"]=mac
    text11.config(text="--",bg="lightgrey")
    text21.config(text="--",bg="lightgrey")
    text31.config(text="--",bg="lightgrey")
    text41.config(text="--",bg="lightgrey")
    text51.config(text="--",bg="lightgrey")
    text61.config(text="--",bg="lightgrey")
    text71.config(text="--",bg="lightgrey")
    text81.config(text="--",bg="lightgrey")
    text91.config(text="--",bg="lightgrey")
    text12.config(text="--",bg="lightgrey")
    text22.config(text="--",bg="lightgrey")
    text32.config(text="--",bg="lightgrey")
    text42.config(text="--",bg="lightgrey")
    text52.config(text="--",bg="lightgrey")
    text62.config(text="--",bg="lightgrey")
    text72.config(text="--",bg="lightgrey")
    text82.config(text="--",bg="lightgrey")
    text92.config(text="--",bg="lightgrey")
    remark_var.set("")

    txt_filename=None

    datee = datetime.now().strftime("%Y-%m-%d")  # Format date as string
    timee = datetime.now().strftime("%H:%M:%S")  # Format time as string
    values["date"]=datee
    values["time"]=timee
    setup_logging(pcb_id)
    with open(txt_filename, mode='w') as tfile:
        tfile.write(f"Date: {datee}\n")
        tfile.write(f"Time: {timee}\n")
        tfile.write(f"Source: {mac}\n")
        tfile.write(f"PCB ID: {pcb_id}\n\n")
    output_area.insert(tk.END, f"Log file created: {txt_filename}\n")

    ser.write("ESP_WDT_START\n".encode('utf-8'))  # Send TESTSTART to serial monitor
    output_area.insert(tk.END, "Sent ESP_WDT_START to serial monitor.\n")

def stop_test():
    global txt_filename
    global trunning
    ser.close()
    updater()
    trunning = False
    label2.config(bg="lightgreen")
    label1.config(bg="white")
    txt_filename=None
    output_area.insert(tk.END, "Logging stopped and file closed.\n")
    ser.flush()
    stop_button.config(state=tk.DISABLED)
    start_button.config(state=tk.NORMAL)
    open_window_button.config(state=tk.NORMAL)

def open_new_window():
    global file_var
    """Open a new window adjacent to the main window with frameS."""
    new_window = tk.Toplevel(app)  # Create a new window
    new_window.title("Flash Tool")  # Set window title
    new_window.geometry("450x100+{}+{}".format(app.winfo_x() + app.winfo_width(), app.winfo_y()))  # Position adjacent to main windo

    # File operation widgets in the new window
    file_var = StringVar()
    frame1=tk.Frame(new_window)
    frame1.pack(pady=5)
    tk.Label(frame1, text="File:").pack(side=tk.LEFT, padx=5)
    tk.Entry(frame1, textvariable=file_var, width=30).pack(side=tk.LEFT, padx=5)
    tk.Button(frame1, text="Browse", command=lambda: browse_file(file_var)).pack(side=tk.LEFT, padx=5)
    frame2=tk.Frame(new_window)
    frame2.pack(pady=5)
    writeb=tk.Button(frame2, text="Write Flash", command=write_flash).pack(side=tk.LEFT, padx=5)
    eraseb=tk.Button(frame2, text="Erase Flash", command=erase_flash).pack(side=tk.LEFT, padx=5)
    readb=tk.Button(frame2, text="Read Flash Info", command=read_flash_info).pack(side=tk.LEFT, padx=5)
    
#----------------------------------------------------------------------------

app = tk.Tk()
app.title("Aquasense PCB QC Test Tool")
app.geometry("760x750")

#----------------------------------------------------------------------------

# Create a frame for the Range section
frame = tk.Frame(app,bd=3, relief="sunken",bg="lightgrey")
frame.pack(pady=5)

# Pressure Sensor in one line
pressure_sensor_label = tk.Label(frame, text="Pressure Sensor - Pressure",bg="lightgrey")
pressure_sensor_label.grid(row=1, column=0, padx=5, pady=5, sticky="e")
pressure_min_entry = tk.Entry(frame, width=3)
pressure_min_entry.grid(row=1, column=1, padx=5)
pressure_min_entry.insert(0,0.9)
hyphen1 = tk.Label(frame, text="-",bg="lightgrey")
hyphen1.grid(row=1, column=2, padx=5)
pressure_max_entry = tk.Entry(frame, width=3)
pressure_max_entry.grid(row=1, column=3, padx=5)
pressure_max_entry.insert(0,1.2)

temperature_label = tk.Label(frame, text="Temperature",bg="lightgrey")
temperature_label.grid(row=1, column=4, padx=5, pady=5, sticky="e")
temp_min_entry = tk.Entry(frame, width=3)
temp_min_entry.grid(row=1, column=5, padx=5)
temp_min_entry.insert(0,5)
hyphen2 = tk.Label(frame, text="-",bg="lightgrey")
hyphen2.grid(row=1, column=6, padx=5)
temp_max_entry = tk.Entry(frame, width=3)
temp_max_entry.grid(row=1, column=7, padx=5)
temp_max_entry.insert(0,25)

# BME Sensor in one line
bme_sensor_label = tk.Label(frame, text="BME Sensor - Pressure",bg="lightgrey")
bme_sensor_label.grid(row=2, column=0, padx=5, pady=5, sticky="e")
bme_pressure_min_entry = tk.Entry(frame, width=3)
bme_pressure_min_entry.grid(row=2, column=1, padx=5)
bme_pressure_min_entry.insert(0,0.9)
hyphen3 = tk.Label(frame, text="-",bg="lightgrey")
hyphen3.grid(row=2, column=2, padx=5)
bme_pressure_max_entry = tk.Entry(frame, width=3)
bme_pressure_max_entry.grid(row=2, column=3, padx=5)
bme_pressure_max_entry.insert(0,1.2)

bme_temperature_label = tk.Label(frame, text="Temperature",bg="lightgrey")
bme_temperature_label.grid(row=2, column=4, padx=5, pady=5, sticky="e")
bme_temp_min_entry = tk.Entry(frame, width=3)
bme_temp_min_entry.grid(row=2, column=5, padx=5)
bme_temp_min_entry.insert(0,5)
hyphen4 = tk.Label(frame, text="-",bg="lightgrey")
hyphen4.grid(row=2, column=6, padx=5)
bme_temp_max_entry = tk.Entry(frame, width=3)
bme_temp_max_entry.grid(row=2, column=7, padx=5)
bme_temp_max_entry.insert(0,25)

bme_altitude_label = tk.Label(frame, text="Altitude",bg="lightgrey")
bme_altitude_label.grid(row=2, column=8, padx=5, pady=5, sticky="e")
bme_altitude_min_entry = tk.Entry(frame, width=3)
bme_altitude_min_entry.grid(row=2, column=9, padx=5)
bme_altitude_min_entry.insert(0,200)
hyphen5 = tk.Label(frame, text="-",bg="lightgrey")
hyphen5.grid(row=2, column=10, padx=5)
bme_altitude_max_entry = tk.Entry(frame, width=3)
bme_altitude_max_entry.grid(row=2, column=11, padx=5)
bme_altitude_max_entry.insert(0,500)

# Battery in one line
battery_label = tk.Label(frame, text="Battery - Voltage",bg="lightgrey")
battery_label.grid(row=3, column=0, padx=5, pady=5, sticky="e")
battery_voltage_min_entry = tk.Entry(frame, width=3)
battery_voltage_min_entry.grid(row=3, column=1, padx=5)
battery_voltage_min_entry.insert(0,3)
hyphen6 = tk.Label(frame, text="-",bg="lightgrey")
hyphen6.grid(row=3, column=2, padx=5)
battery_voltage_max_entry = tk.Entry(frame, width=3)
battery_voltage_max_entry.grid(row=3, column=3, padx=5)
battery_voltage_max_entry.insert(0,3.5)


#----------------------------------------------------------------------------

port_var = StringVar()
port_frame = tk.Frame(app)
port_frame.pack(pady=5)
tk.Label(port_frame, text="Select Com Port - ",font=("Poppins", 12, "bold")).pack(side=tk.LEFT, padx=5)
port_combobox = ttk.Combobox(port_frame, textvariable=port_var, values=get_ports())
port_combobox.pack(side=tk.LEFT)

connect_button = tk.Button(port_frame, text="Connect", command=connect)
connect_button.pack(side=tk.LEFT, padx=5)

refresh_button = tk.Button(port_frame, text="Refresh Ports", command=refresh_ports)
refresh_button.pack(side=tk.LEFT, padx=5)

open_window_button = tk.Button(port_frame, text="Open Flash Tool ->", command=open_new_window)
open_window_button.pack(side=tk.LEFT,pady=10)

#----------------------------------------------------------------------------

frame0=tk.Frame(app)
pcbid_var = StringVar()
frame0.pack(pady=5)
pcbid_var.trace_add("write", enable_start_button)  # Trigger enable button on text change
tk.Label(frame0, text="PCB ID - ",font=("Poppins", 12, "bold")).pack(side=tk.LEFT)
tk.Entry(frame0, textvariable=pcbid_var, width=20).pack(side=tk.LEFT, padx=5)

start_button = tk.Button(frame0, text="Start", command=start_test,state=tk.DISABLED)
start_button.pack(side=tk.LEFT, padx=5)

stop_button = tk.Button(frame0, text="Stop", command=stop_test,state=tk.DISABLED)
stop_button.pack(side=tk.LEFT, padx=5)

so_button = tk.Button(frame0, text="Start over",command=startover)
so_button.pack(side=tk.LEFT, padx=5)

#----------------------------------------------------------------------------

frameS=tk.Frame(app)
frameS.pack(pady=5)

tr=ts=tk.Label(frameS,text="Running",font=("Poppins", 12, "bold"))
tr.pack(side=tk.LEFT)
label1 = tk.Label(frameS, width=2, height=1, relief="solid", bg="white")
label1.pack(side=tk.LEFT,padx=5)
tc=ts=tk.Label(frameS,text="QC Done",font=("Poppins", 12, "bold"))
tc.pack(side=tk.LEFT)
label2 = tk.Label(frameS, width=2, height=1, relief="solid", bg="white")
label2.pack(side=tk.LEFT)
time_label = tk.Label(frameS, text="Timer: 00:00", font=("Poppins", 12, "bold"))
time_label.pack(side=tk.LEFT,padx=5)

frame1 = tk.Frame(app,bd=3, relief="sunken",bg="lightgrey")
frame1.pack(fill="both",pady=5, padx=10)

#headings
text00 = tk.Label(frame1, text="Test",font=("Poppins", 14, "bold"),bg="grey",anchor="w")
text00.grid(row=0, column=0, sticky="nsew")
text01 = tk.Label(frame1, text="State",font=("Poppins", 14, "bold"),bg="grey")
text01.grid(row=0, column=1, sticky="nsew")
text02 = tk.Label(frame1, text="Accept/Reject",font=("Poppins", 14, "bold"),bg="grey")
text02.grid(row=0, column=2, sticky="nsew")
tk.Label(frame1, text="",font=("Poppins", 14, "bold"),bg="grey").grid(row=0, column=3, sticky="nsew")
tk.Label(frame1, text="",font=("Poppins", 14, "bold"),bg="grey").grid(row=0, column=4, sticky="nsew")

#Watch Dog Timer
text10 = tk.Label(frame1, text="Watch Dog Timer",anchor="w",bg="lightgrey")
text10.grid(row=1, column=0,sticky="nsew")
text11 = tk.Label(frame1, text="--",bg="lightgrey")
text11.grid(row=1, column=1,sticky="nsew")
text12 = tk.Label(frame1, text="--",bg="lightgrey")
text12.grid(row=1, column=2,sticky="nsew")
b13=tk.Button(frame1,text="Accept",highlightbackground="lightgrey",command=wdt_accept)
b13.grid(row=1,column=3)
b14=tk.Button(frame1,text="Reject",highlightbackground="lightgrey",command=wdt_reject)
b14.grid(row=1,column=4)

#ESP32 Flash Info
text20 = tk.Label(frame1, text="ESP32 Flash Info",anchor="w",bg="lightgrey")
text20.grid(row=2, column=0,sticky="nsew")
text21 = tk.Label(frame1, text="--",bg="lightgrey")
text21.grid(row=2, column=1,sticky="nsew")
text22 = tk.Label(frame1, text="--",bg="lightgrey")
text22.grid(row=2, column=2,sticky="nsew")
b23=tk.Button(frame1,text="Accept",highlightbackground="lightgrey",command=info_accept)
b23.grid(row=2,column=3)
b24=tk.Button(frame1,text="Reject",highlightbackground="lightgrey",command=info_reject)
b24.grid(row=2,column=4)

#RTC Communication
text30 = tk.Label(frame1, text="RTC Communication",anchor="w",bg="lightgrey")
text30.grid(row=3, column=0,sticky="nsew")
text31 = tk.Label(frame1, text="--",bg="lightgrey")
text31.grid(row=3, column=1,sticky="nsew")
text32 = tk.Label(frame1, text="--",bg="lightgrey")
text32.grid(row=3, column=2,sticky="nsew")
b33=tk.Button(frame1,text="Accept",highlightbackground="lightgrey",command=rtc_accept)
b33.grid(row=3,column=3)
b34=tk.Button(frame1,text="Reject",highlightbackground="lightgrey",command=rtc_reject)
b34.grid(row=3,column=4)

#Pressure Sensor Communication
text40 = tk.Label(frame1, text="Pressure Sensor Communication",anchor="w",bg="lightgrey")
text40.grid(row=4, column=0,sticky="nsew")
text41 = tk.Label(frame1, text="--",bg="lightgrey")
text41.grid(row=4, column=1,sticky="nsew")
text42 = tk.Label(frame1, text="--",bg="lightgrey")
text42.grid(row=4, column=2,sticky="nsew")
b43=tk.Button(frame1,text="Accept",highlightbackground="lightgrey",command=pressure_accept)
b43.grid(row=4,column=3)
b44=tk.Button(frame1,text="Reject",highlightbackground="lightgrey",command=pressure_reject)
b44.grid(row=4,column=4)

#Battery Voltage
text50 = tk.Label(frame1, text="Battery Voltage",anchor="w",bg="lightgrey")
text50.grid(row=5, column=0,sticky="nsew")
text51 = tk.Label(frame1, text="--",bg="lightgrey")
text51.grid(row=5, column=1,sticky="nsew")
text52 = tk.Label(frame1, text="--",bg="lightgrey")
text52.grid(row=5, column=2,sticky="nsew")
b53=tk.Button(frame1,text="Accept",highlightbackground="lightgrey",command=battery_accept)
b53.grid(row=5,column=3)
b54=tk.Button(frame1,text="Reject",highlightbackground="lightgrey",command=battery_reject)
b54.grid(row=5,column=4)

#SD Card Read+Write
text60 = tk.Label(frame1, text="SD Card Read+Write",anchor="w",bg="lightgrey")
text60.grid(row=6, column=0,sticky="nsew")
text61 = tk.Label(frame1, text="--",bg="lightgrey")
text61.grid(row=6, column=1,sticky="nsew")
text62 = tk.Label(frame1, text="--",bg="lightgrey")
text62.grid(row=6, column=2,sticky="nsew")
b63=tk.Button(frame1,text="Accept",highlightbackground="lightgrey",command=sd_accept)
b63.grid(row=6,column=3)
b64=tk.Button(frame1,text="Reject",highlightbackground="lightgrey",command=sd_reject)
b64.grid(row=6,column=4)

#BME Sensor Communication
text70 = tk.Label(frame1, text="BME Sensor Communication",anchor="w",bg="lightgrey")
text70.grid(row=7, column=0,sticky="nsew")
text71 = tk.Label(frame1, text="--",bg="lightgrey")
text71.grid(row=7, column=1,sticky="nsew")
text72 = tk.Label(frame1, text="--",bg="lightgrey")
text72.grid(row=7, column=2,sticky="nsew")
b73=tk.Button(frame1,text="Accept",highlightbackground="lightgrey",command=bme_accept)
b73.grid(row=7,column=3)
b74=tk.Button(frame1,text="Reject",highlightbackground="lightgrey",command=bme_reject)
b74.grid(row=7,column=4)

#Network Connection
text80 = tk.Label(frame1, text="GPRS Communication",anchor="w",bg="lightgrey")
text80.grid(row=8, column=0,sticky="nsew")
text81 = tk.Label(frame1, text="--",bg="lightgrey")
text81.grid(row=8, column=1,sticky="nsew")
text82 = tk.Label(frame1, text="--",bg="lightgrey")
text82.grid(row=8, column=2,sticky="nsew")
b83=tk.Button(frame1,text="Accept",highlightbackground="lightgrey",command=net_accept)
b83.grid(row=8,column=3)
b84=tk.Button(frame1,text="Reject",highlightbackground="lightgrey",command=net_reject)
b84.grid(row=8,column=4)

#hibernate
text90 = tk.Label(frame1, text="Hibernate Mode Testing",anchor="w",bg="lightgrey")
text90.grid(row=9, column=0,sticky="nsew")
text91 = tk.Label(frame1, text="--",bg="lightgrey")
text91.grid(row=9, column=1,sticky="nsew")
text92 = tk.Label(frame1, text="--",bg="lightgrey")
text92.grid(row=9, column=2,sticky="nsew")
b93=tk.Button(frame1,text="Accept",highlightbackground="lightgrey",command=hibernate_accept)
b93.grid(row=9,column=3)
b94=tk.Button(frame1,text="Reject",highlightbackground="lightgrey",command=hibernate_reject)
b94.grid(row=9,column=4)

frame1.grid_columnconfigure(0, weight=7)  
frame1.grid_columnconfigure(1, weight=1)  
frame1.grid_columnconfigure(2, weight=1)  
frame1.grid_columnconfigure(3, weight=1) 

#----------------------------------------------------------------------------

remark_var = StringVar()
remark_frame = tk.Frame(app)
remark_frame.pack(pady=5)
tk.Label(remark_frame, text="Remark - ",font=("Poppins", 12, "bold")).pack(side=tk.LEFT)
tk.Entry(remark_frame, textvariable=remark_var, width=30).pack(side=tk.LEFT, padx=5)

#----------------------------------------------------------------------------

label_frame = tk.Frame(app)
label_frame.pack(padx=10, pady=5)
label4 = tk.Button(
    label_frame, text="Pass",bg="lightblue",command=passtest,height=3,width=20)
label4.pack(side=tk.LEFT,pady=5)
label5 = tk.Button(
    label_frame, text="Fail",command=failtest,height=3,width=20)
label5.pack(side=tk.LEFT,pady=5)
#----------------------------------------------------------------------------
output_area = tk.Text(app, height=7, width=100, state=tk.NORMAL)
output_area.pack()
values={"wdt":"--","info":"--","rtc":"--","pressure":"--","battery":"--","sd":"--","bme":"--","net":"--","hibernate":"--"}
ser = None
is_port_running = False
hflag=None
txt_filename=None
datee=None
switchflag=True
timer_flag = False
timee=None
wdtflag=False
qcflag=None
file_var=None
trunning = False
counter = 0  
mac=None

refresh_ports()

app.mainloop()


