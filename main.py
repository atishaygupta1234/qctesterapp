import tkinter as tk
from tkinter import StringVar, ttk
import serial
import serial.tools.list_ports
import threading
from datetime import datetime
import logging
from tkinter import ttk
import uuid
from openpyxl import Workbook, load_workbook
import os
import time

def updater():
    global values
    pmin=float(pressure_min_entry.get())
    values["pmin"]=pmin
    pmax=float(pressure_max_entry.get())
    values["pmax"]=pmax
    tmin=float(temp_min_entry.get())
    values["tmin"]=tmin
    tmax=float(temp_max_entry.get())
    values["tmax"]=tmax
    bpmin=float(bme_pressure_min_entry.get())
    values["bpmin"]=bpmin
    bpmax=float(bme_pressure_max_entry.get())
    values["bpmax"]=bpmax
    btmin=float(bme_temp_min_entry.get())
    values["btmin"]=btmin
    btmax=float(bme_temp_max_entry.get())
    values["btmax"]=btmax
    bamin=float(bme_altitude_min_entry.get())
    values["bamin"]=bamin
    bamax=float(bme_altitude_max_entry.get())
    values["bamax"]=bamax
    vmin=float(battery_voltage_min_entry.get())
    values["vmin"]=vmin
    vmax=float(battery_voltage_max_entry.get())
    values["vmax"]=vmax
    totaltime=time_label.cget("text")[7:]
    values["totaltime"]=totaltime
    print(values)

    if(values["wdt"]=="PASS"):
        wdt_accept()
        values["finalwdt"]="PASS"
    else:
        wdt_reject()
        values["finalwdt"]="FAIL"

    if(values["info"]!="--"):
        info_accept()
        values["espimei"]=values["info"][0]
        values["espsize"]=values["info"][1]
        values["espmac"]=values["info"][2]
        values["finalinfo"]="PASS"
    else:
        info_reject()
        values["finalinfo"]="FAIL"
        values["espimei"]="--"
        values["espsize"]="--"
        values["espmac"]="--"

    if(values["rtc"]=="PASS"):
        rtc_accept()
        values["finalrtc"]="PASS"
    else:
        rtc_reject()
        values["finalrtc"]="FAIL"


    if(values["pressure"]=="FAIL"): 
        pressure_reject()
        values["ppressure"]="--"
        values["ptemp"]="--"
        values["finalpressure"]="FAIL"
    else:
        if(values["pressure"]!="--"):
            values["ppressure"]=float(values["pressure"][0])
            values["ptemp"]=float(values["pressure"][1])
            if((pmin<=values["ppressure"]<=pmax) and (tmin<=values["ptemp"]<=tmax)):
                values["finalpressure"]="PASS"
                pressure_accept()
            else:
                values["finalpressure"]="FAIL"
                pressure_reject()
        else:
            pressure_reject()
            values["ppressure"]="--"
            values["ptemp"]="--"
            values["finalpressure"]="FAIL"

    if(values["battery"]=="FAIL"): 
        battery_reject()
        values["finalbattery"]="FAIL"
        values["voltage"]="--"
    else:
        if(values["battery"]!="--"):
            values["voltage"]=float(values["battery"][0])
            if(vmin<=float(values["voltage"])<=vmax): 
                values["finalbattery"]="PASS"
                battery_accept()
            else:
                battery_reject()
                values["finalbattery"]="FAIL"
        else:
            battery_reject()
            values["finalbattery"]="FAIL"
            values["voltage"]="--"

    if(values["sd"]=="PASS"):
        sd_accept()
        values["finalsd"]="PASS"
    else:
        sd_reject()
        values["finalsd"]="FAIL"

    if(values["bme"]=="FAIL"): 
        bme_reject()
        values["bpressure"]="--"
        values["btemp"]="--"
        values["baltitude"]="--"
        values["finalbme"]="FAIL"
    else:
        if(values["bme"]!="--"):
            values["bpressure"]=float(values["bme"][0])
            values["btemp"]=float(values["bme"][1])
            values["baltitude"]=float(values["bme"][2])
            if((bpmin<=values["bpressure"]<=bpmax)and(btmin<=values["btemp"]<=btmax)and(bamin<=values["baltitude"]<=bamax)):
                values["finalbme"]="PASS"
                bme_accept()
            else:
                values["finalbme"]="FAIL"
                bme_reject()
        else:
            bme_reject()
            values["bpressure"]="--"
            values["btemp"]="--"
            values["baltitude"]="--"
            values["finalbme"]="FAIL"

    if(values["net"]=="PASS"):
        net_accept()
        values["finalnet"]="PASS"
    else:
        net_reject()
        values["finalnet"]="FAIL"

    if(values["hibernate"]=="PASS"):
        hibernate_accept()
        values["finalhib"]="PASS"
    else:
        hibernate_reject()
        values["finalhib"]="FAIL"

    print(values)

def wdt_accept():
    text10.config(bg="lightgreen")
    text11.config(bg="lightgreen")
    text12.config(bg="lightgreen",text="Accepted")
    
def wdt_reject():
    text10.config(bg="lightcoral")
    text11.config(bg="lightcoral")
    text12.config(bg="lightcoral",text="Rejected")

def info_accept():
    text20.config(bg="lightgreen")
    text21.config(bg="lightgreen")
    text22.config(bg="lightgreen",text="Accepted")

def info_reject():
    text20.config(bg="lightcoral")
    text21.config(bg="lightcoral")
    text22.config(bg="lightcoral",text="Rejected")

def rtc_accept():
    text30.config(bg="lightgreen")
    text31.config(bg="lightgreen")
    text32.config(bg="lightgreen",text="Accepted")

def rtc_reject():
    text30.config(bg="lightcoral")
    text31.config(bg="lightcoral")
    text32.config(bg="lightcoral",text="Rejected")

def pressure_accept():
    text40.config(bg="lightgreen")
    text41.config(bg="lightgreen")
    text42.config(bg="lightgreen",text="Accepted")

def pressure_reject():
    text40.config(bg="lightcoral")
    text41.config(bg="lightcoral")
    text42.config(bg="lightcoral",text="Rejected")

def battery_accept():
    text50.config(bg="lightgreen")
    text51.config(bg="lightgreen")
    text52.config(bg="lightgreen",text="Accepted")

def battery_reject():
    text50.config(bg="lightcoral")
    text51.config(bg="lightcoral")
    text52.config(bg="lightcoral",text="Rejected")

def sd_accept():
    text60.config(bg="lightgreen")
    text61.config(bg="lightgreen")
    text62.config(bg="lightgreen",text="Accepted")

def sd_reject():
    text60.config(bg="lightcoral")
    text61.config(bg="lightcoral")
    text62.config(bg="lightcoral",text="Rejected")

def bme_accept():
    text70.config(bg="lightgreen")
    text71.config(bg="lightgreen")
    text72.config(bg="lightgreen",text="Accepted")

def bme_reject():
    text70.config(bg="lightcoral")
    text71.config(bg="lightcoral")
    text72.config(bg="lightcoral",text="Rejected")

def net_accept():
    text80.config(bg="lightgreen")
    text81.config(bg="lightgreen")
    text82.config(bg="lightgreen",text="Accepted")

def net_reject():
    text80.config(bg="lightcoral")
    text81.config(bg="lightcoral")
    text82.config(bg="lightcoral",text="Rejected")

def hibernate_accept():
    text90.config(bg="lightgreen")
    text91.config(bg="lightgreen")
    text92.config(bg="lightgreen",text="Accepted")

def hibernate_reject():
    text90.config(bg="lightcoral")
    text91.config(bg="lightcoral")
    text92.config(bg="lightcoral",text="Rejected")

def update_timer():
    global trunning, counter
    while trunning:
        minutes, seconds = divmod(counter, 60)
        timer_text = f"Timer: {minutes:02}:{seconds:02}"
        time_label.config(text=timer_text)
        app.update()
        counter += 1
        threading.Event().wait(1)

def get_mac_address():
    mac = uuid.getnode()
    mac_address = ':'.join(f'{(mac >> i) & 0xff:02x}' for i in range(0, 8 * 6, 8)[::-1])
    return mac_address

def passtest():
    fail_label.config(state=tk.DISABLED)
    pass_label.config(bg="lightgreen")
    file_path = "output.xlsx"
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Date", "Time", "Pcb Id","Source Id","ESP32 IMEI","ESP32 SIZE","ESP32 Info Communication","Watch Dog Timer","RTC Communication","PT - Pressure","PT - Temperature","Pressure Sensor Communication","Battery Voltage","Battery Communication","SD Card Communication","BME - Pressure","BME - Temperature","BME -Altitude","BME Sensor Communcation","GPRS Communication","Hibernate Mode Testing","Comments","Time Taken","Final Status"])
    values['ffinalinfo'] = "Pass" if text22.cget("text") == "Accepted" else "Fail"
    values['ffinalbattery'] = "Pass" if text52.cget("text") == "Accepted" else "Fail"
    values['ffinalbme'] = "Pass" if text72.cget("text") == "Accepted" else "Fail"
    values["ffinalhib"] = "Pass" if text92.cget("text") == "Accepted" else "Fail"
    values["ffinalwdt"] = "Pass" if text12.cget("text") == "Accepted" else "Fail"
    values["ffinalrtc"] = "Pass" if text32.cget("text") == "Accepted" else "Fail"
    values["ffinalpressure"] = "Pass" if text42.cget("text") == "Accepted" else "Fail"
    values["ffinalsd"] = "Pass" if text62.cget("text") == "Accepted" else "Fail"
    values["ffinalnet"] = "Pass" if text82.cget("text") == "Accepted" else "Fail"
    row_to_insert = [values.get("date"), values.get("time"), values.get("pcbid"),values.get("sourceid"),values.get("espimei"),values.get("espsize"),values.get("ffinalinfo"),values.get("ffinalwdt"),values.get("ffinalrtc"),values.get("ppressure"),values.get("ptemp"),values.get("ffinalpressure"),values.get("voltage"),values.get("ffinalbattery"),values.get("ffinalsd"),values.get("bpressure"),values.get("btemp"),values.get("baltitude"),values.get("ffinalbme"),values.get("ffinalnet"),values.get("ffinalhib"),remark_var.get(),values.get("totaltime"),"PASS"]
    sheet.append(row_to_insert)
    wb.save(file_path)
    output_area.insert(tk.END, f"Data successfully written to {file_path}.\n")
    print(f"Data successfully written to {file_path}.")

def failtest():
    pass_label.config(state=tk.DISABLED)
    fail_label.config(bg="lightcoral")
    file_path = "output.xlsx"
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Date", "Time", "Pcb Id","Source Id","ESP32 IMEI","ESP32 SIZE","ESP32 Info Communication","Watch Dog Timer","RTC Communication","PT - Pressure","PT - Temperature","Pressure Sensor Communication","Battery Voltage","Battery Communication","SD Card Communication","BME - Pressure","BME - Temperature","BME - Altitude","BME Sensor Communcation","GPRS Communication","Hibernate Mode Testing","Comments","Time Taken","Final Status"])
    values['ffinalinfo'] = "Pass" if text22.cget("text") == "Accepted" else "Fail"
    values['ffinalbattery'] = "Pass" if text52.cget("text") == "Accepted" else "Fail"
    values['ffinalbme'] = "Pass" if text72.cget("text") == "Accepted" else "Fail"
    values["ffinalhib"] = "Pass" if text92.cget("text") == "Accepted" else "Fail"
    values["ffinalwdt"] = "Pass" if text12.cget("text") == "Accepted" else "Fail"
    values["ffinalrtc"] = "Pass" if text32.cget("text") == "Accepted" else "Fail"
    values["ffinalpressure"] = "Pass" if text42.cget("text") == "Accepted" else "Fail"
    values["ffinalsd"] = "Pass" if text62.cget("text") == "Accepted" else "Fail"
    values["ffinalnet"] = "Pass" if text82.cget("text") == "Accepted" else "Fail"
    row_to_insert = [values.get("date"), values.get("time"), values.get("pcbid"),values.get("sourceid"),values.get("espimei"),values.get("espsize"),values.get("ffinalinfo"),values.get("ffinalwdt"),values.get("ffinalrtc"),values.get("ppressure"),values.get("ptemp"),values.get("ffinalpressure"),values.get("voltage"),values.get("ffinalbattery"),values.get("ffinalsd"),values.get("bpressure"),values.get("btemp"),values.get("baltitude"),values.get("ffinalbme"),values.get("ffinalnet"),values.get("ffinalhib"),remark_var.get(),values.get("totaltime"),"FAIL"]
    sheet.append(row_to_insert)
    wb.save(file_path)
    output_area.insert(tk.END, f"Data successfully written to {file_path}.\n")
    print(f"Data successfully written to {file_path}.")

def setup_logging(pcbid):
    global txt_filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    txt_filename = f"{pcbid}_{timestamp}.txt"

    logging.basicConfig(
        filename=txt_filename,
        level=logging.DEBUG,  # You can change the level to INFO or ERROR depending on what you need
        format="%(asctime)s - %(message)s",  # You can adjust the format as needed
    )

def enable_start_button(*args):
    if (pcbid_var.get() and is_port_running ==True):
        start_button.config(state=tk.NORMAL)
        stop_button.config(state=tk.DISABLED)

def get_ports():
    return [port.device for port in serial.tools.list_ports.comports()]

def refresh_port1():
    port_combobox1['values'] = get_ports()
    if port_combobox1['values']:
        port_combobox1.current(0)

def refresh_port2():
    port_combobox2['values'] = get_ports()
    if port_combobox2['values']:
        port_combobox2.current(0)

def connect():
    global ser, is_port_running
    port = port_var1.get()

    if not port:
        output_area.insert(tk.END, "Please select a COM port.\n")
        return
    
    output_area.insert(tk.END, "Connecting...\n")

    try:
        ser = serial.Serial(port, 115200, timeout=1)
        ser.flushInput()
        is_port_running = True
        output_area.insert(tk.END, f"Connected to {port} at 115200 baud.\n")
        threading.Thread(target=read_from_serial, daemon=True).start()
    except Exception as e:
        output_area.insert(tk.END, f"Failed to connect: {e}\n")
        is_port_running = False

def startover():
    global is_port_running,ser,datee,timee,values,hflag,completeflag,trunning, counter,ser2,mac,txt_filename,intime,wdtflag,qcflag,data
    trunning = False
    txt_filename=None 
    intime=True 
    qcflag=None 
    ser2=None 
    data=None
    counter = 0
    time_label.config(text="Timer: 00:00")
    pass_label.config(state=tk.NORMAL)
    fail_label.config(state=tk.NORMAL)
    pass_label.config(fg="black")
    fail_label.config(fg="black")
    is_port_running = False
    ser = None
    datee=None
    timee=None
    hflag=None
    mac=None
    values={"wdt":"--","info":"--","rtc":"--","pressure":"--","battery":"--","sd":"--","bme":"--","net":"--","hibernate":"--"}
    start_button.config(state=tk.DISABLED)
    stop_button.config(state=tk.DISABLED)
    pcbid_var.set("")
    remark_var.set("")
    completeflag=False
    running_label.config(bg="white")
    qcdone_label.config(bg="white")
    stopped_label.config(bg="white")
    pass_label.config(bg="white")
    fail_label.config(bg="white")
    text10.config(bg="lightgrey")
    text20.config(bg="lightgrey")
    text30.config(bg="lightgrey")
    text40.config(bg="lightgrey")
    text50.config(bg="lightgrey")
    text60.config(bg="lightgrey")
    text70.config(bg="lightgrey")
    text80.config(bg="lightgrey")
    text90.config(bg="lightgrey")
    text11.config(text="--",bg="lightgrey")
    text21.config(text="--",bg="lightgrey")
    text31.config(text="--",bg="lightgrey")
    text41.config(text="--",bg="lightgrey")
    text51.config(text="--",bg="lightgrey")
    text61.config(text="--",bg="lightgrey")
    text71.config(text="--",bg="lightgrey")
    text81.config(text="--",bg="lightgrey")
    text91.config(text="--",bg="lightgrey")
    text12.config(text="--",bg="lightgrey")
    text22.config(text="--",bg="lightgrey")
    text32.config(text="--",bg="lightgrey")
    text42.config(text="--",bg="lightgrey")
    text52.config(text="--",bg="lightgrey")
    text62.config(text="--",bg="lightgrey")
    text72.config(text="--",bg="lightgrey")
    text82.config(text="--",bg="lightgrey")
    text92.config(text="--",bg="lightgrey")

def read_from_serial():
    global ser, is_port_running,wdtflag,qcflag,hflag,values,completeflag,data
    while is_port_running:
        try:
            if ser.in_waiting > 0:
                data = ser.readline().decode('utf-8', errors='ignore').strip()
                output_area.insert(tk.END, f"{data}\n")
                output_area.see(tk.END)  # Auto-scroll to the bottom

                if data == "[ExC: Main] ESP_WDT_START_ACK":  
                    output_area.insert(tk.END, "Received ESP_WDT_START_ACK.\n")

                if data.startswith("[ExC: Main] WDT:"):
                    logging.info("WDT Test Starts...")
                    wdtflag=True

                if data.startswith("[ExC Main] QC Firmware") and wdtflag==True:
                    logging.info("[ExC: Main] WDT: Pass")
                    values["wdt"]="PASS"
                    text11.config(text="PASS",bg="lightgrey")
                    wdtflag=False
                    ser.write("ESP_QC_START\n".encode('utf-8'))  

                if wdtflag==True and data.startswith("[ExC: Main] WDT: Failed"):
                    logging.info("[ExC: Main] WDT: Fail")
                    text11.config(text="Fail",bg="lightgrey")
                    values["wdt"]="FAIL"
                    wdtflag=False
                    ser.write("ESP_QC_START\n".encode('utf-8')) 

                if data.startswith("[ExC: Main] ESP_START_ACK"):
                    qcflag=True
                    pfail=False
                    bfail=False
                    logging.info("QC Tests Starts...")

                if data.startswith("[ExC: Main] ESPinfo:") and qcflag==True:
                    oldvals=data[21:].strip()
                    vals= oldvals.split(',')
                    values["info"]=vals
                    text21.config(text=f"{oldvals}",bg="lightgrey")
                    logging.info(f"{data}")

                if data.startswith("[ExC: Main] RTCInterface: Pass") and qcflag==True:
                    text31.config(text="Pass",bg="lightgrey")
                    logging.info(f"{data}")
                    values["rtc"]="PASS"

                if data.startswith("[ExC: Main] RTCInterface: Failed") and qcflag==True:
                    text31.config(text="Fail",bg="lightgrey")
                    logging.info(f"{data}")
                    values["rtc"]="FAIL"

                if data=="[ExC: Main] PressureTransmitter: Failed" and qcflag==True:
                    text41.config(text="Fail",bg="lightgrey")
                    logging.info(f"{data}")
                    values["pressure"]="FAIL"
                    pfail=True

                if data.startswith("[ExC: Main] PressureTransmitter:") and qcflag==True and pfail==False:
                    oldvals=data[33:].strip()
                    vals= oldvals.split(',')
                    text41.config(text=f"{oldvals}",bg="lightgrey")
                    logging.info(f"{data}")
                    values["pressure"]=vals
                    pfail=True

                if data.startswith("[ExC: Main] BatteryVoltage:") and qcflag==True:
                    oldvals=data[28:].strip()
                    vals= oldvals.split(',')
                    values["battery"]=vals
                    text51.config(text=f"{oldvals}",bg="lightgrey")
                    logging.info(f"{data}")

                if data == "[ExC: Main] uSD: Pass" and qcflag==True:
                    text61.config(text="Pass",bg="lightgrey")
                    values["sd"]="PASS"
                    logging.info(f"{data}")

                if data == "[ExC: Main] uSD: Failed" and qcflag==True:
                    text61.config(text="Fail",bg="lightgrey")
                    values["sd"]="FAIL"
                    logging.info(f"{data}")

                if data=="[ExC: Main] BME: Failed" and qcflag==True:
                    text71.config(text="Fail",bg="lightgrey")
                    logging.info(f"{data}")
                    values["bme"]="FAIL"
                    bfail=True

                if data.startswith("[ExC: Main] BME:") and qcflag==True and bfail==False:
                    oldvals=data[17:].strip()
                    vals= oldvals.split(',')
                    values["bme"]=vals
                    text71.config(text=f"{oldvals}",bg="lightgrey")
                    logging.info(f"{data}")
                    bfail=True

                if data == "[ExC: Main] GPRSConnection: Pass" and qcflag==True:
                    text81.config(text="Pass",bg="lightgrey")
                    values["net"]="PASS"
                    logging.info(f"{data}")
                    qcflag=False
                    hflag=True
                    ser.write("ESP_LPM_START\n".encode('utf-8')) 

                if data == "[ExC: Main] GPRSConnection: Failed" and qcflag==True:
                    text81.config(text="Fail",bg="lightgrey")
                    values["net"]="FAIL"
                    logging.info(f"{data}")
                    qcflag=False
                    hflag=True
                    ser.write("ESP_LPM_START\n".encode('utf-8')) 

                if data == "[ExC: Main] ESP_LPM_ACK" and hflag == True:
                    output_area.insert(tk.END, "LPM acknowledged... starting timer\n")
                    start_timer()
                    handle_lpm_ack()  # Start periodic check for QC Firmware
                
        except Exception as e:
            output_area.insert(tk.END, f"Error reading from serial: {e}\n")
            output_area.see(tk.END)
            break

def handle_lpm_ack():
    global intime, hflag, completeflag

    if intime:  # Timer is running
        if data == "[ExC Main] QC Firmware" and hflag == True:
            text91.config(text="Pass", bg="lightgrey")
            logging.info(f"LPM Pass")
            values["hibernate"] = "PASS"
            hflag = False
            completeflag = True
            stop_test()
        else:
            # Re-check after 1 second
            app.after(1000, handle_lpm_ack)
    else:  # Timer expired
        output_area.insert(tk.END, "Timer expired. No QC Firmware message received.\n")
        values["hibernate"]="FAIL"
        hflag = False
        completeflag = True
        stop_test()

def start_timer():
    global intime
    print("timer started")
    def timer_function():
        time.sleep(120)  # Wait for 2 minutes (120 seconds)
        global intime
        intime = False
        print("Timer finished. intime set to False.")
    
    # Start the timer in a separate thread
    timer_thread = threading.Thread(target=timer_function)
    timer_thread.start()

def start_test():
    global ser,datee,timee,txt_filename,mac
    global trunning
    if not trunning:
        trunning = True
        threading.Thread(target=update_timer, daemon=True).start()
    running_label.config(bg="lightgreen")
    start_button.config(state=tk.DISABLED)
    stop_button.config(state=tk.NORMAL)
    if not ser or not ser.is_open:
        output_area.insert(tk.END, "No active serial connection. Startover and Connect to a port first.\n")
        return
    pcb_id = pcbid_var.get()
    mac= get_mac_address()
    values["pcbid"]=pcb_id
    values["sourceid"]=mac
    '''text11.config(text="--",bg="lightgrey")
    text21.config(text="--",bg="lightgrey")
    text31.config(text="--",bg="lightgrey")
    text41.config(text="--",bg="lightgrey")
    text51.config(text="--",bg="lightgrey")
    text61.config(text="--",bg="lightgrey")
    text71.config(text="--",bg="lightgrey")
    text81.config(text="--",bg="lightgrey")
    text91.config(text="--",bg="lightgrey")
    text12.config(text="--",bg="lightgrey")
    text22.config(text="--",bg="lightgrey")
    text32.config(text="--",bg="lightgrey")
    text42.config(text="--",bg="lightgrey")
    text52.config(text="--",bg="lightgrey")
    text62.config(text="--",bg="lightgrey")
    text72.config(text="--",bg="lightgrey")
    text82.config(text="--",bg="lightgrey")
    text92.config(text="--",bg="lightgrey")'''
    remark_var.set("")
    datee = datetime.now().strftime("%Y-%m-%d") 
    timee = datetime.now().strftime("%H:%M:%S")  
    values["date"]=datee
    values["time"]=timee
    setup_logging(pcb_id)
    with open(txt_filename, mode='w') as tfile:
        tfile.write(f"Date: {datee}\n")
        tfile.write(f"Time: {timee}\n")
        tfile.write(f"Source: {mac}\n")
        tfile.write(f"PCB ID: {pcb_id}\n\n")
    output_area.insert(tk.END, f"Log file created: {txt_filename}\n")

    ser.write("ESP_WDT_START\n".encode('utf-8')) 
    output_area.insert(tk.END, "Sent ESP_WDT_START to serial monitor.\n")

def stop_test():
    global trunning,ser
    updater()
    trunning = False
    running_label.config(bg="white")
    if completeflag==True:
        print("complete hoke aya hu")
        qcdone_label.config(bg="lightgreen")
    else:
        print("zabardasti agya")
        stopped_label.config(bg="red")
    output_area.insert(tk.END, "Logging stopped and file closed.\n")
    ser.flush()
    ser.close()
    stop_button.config(state=tk.DISABLED)
    start_button.config(state=tk.NORMAL)

def connect2():
    global ser2
    port = port_var2.get()  # Get selected port from Combobox
    if not port:
        output_area.insert(tk.END, "No port selected. Please select a port.\n")
        return

    try:
        ser2 = serial.Serial(port, baudrate=115200, timeout=1)
        output_area.insert(tk.END, f"Connected to {port}. Waiting for command...\n")

        # Start a thread to listen for the specific command
        threading.Thread(target=listen_for_command, daemon=True).start()

    except Exception as e:
        output_area.insert(tk.END, f"Failed to connect to {port}. Error: {e}\n")

def listen_for_command():
    global ser2
    try:
        while ser2 and ser2.is_open:
            line = ser2.readline().decode('utf-8', errors='ignore').strip()
            if line:  # If any data is received
                output_area.insert(tk.END, f"{line}\n")
                output_area.see(tk.END)  # Auto-scroll to the end
                
                if "[ExC Main] PSUP_ON" in line:
                    output_area.insert(tk.END, "Switch to PCB Supply, remove USB-C and connect the debug cable\n")
                    output_area.see(tk.END)
                    disconnect2()  # Move to disconnect function
                    break
            time.sleep(0.1)  # Avoid high CPU usage

    except Exception as e:
        output_area.insert(tk.END, f"Error while reading: {e}\n")

def disconnect2():
    global ser2
    if ser2 and ser2.is_open:
        ser2.close()
        output_area.insert(tk.END, "Disconnected from port.\n")
    else:
        output_area.insert(tk.END, "No active connection on Debug Port to disconnect.\n")

#----------------------------------------------------------------------------

app = tk.Tk()
app.title("Aquasense PCB QC Test Tool")
app.geometry("760x750")

frame = tk.Frame(app,bd=3, relief="sunken",bg="lightgrey")
frame.pack(pady=5)

pressure_sensor_label = tk.Label(frame, text="Pressure Sensor - Pressure",bg="lightgrey")
pressure_sensor_label.grid(row=1, column=0, padx=5, pady=5, sticky="e")
pressure_min_entry = tk.Entry(frame, width=3)
pressure_min_entry.grid(row=1, column=1, padx=5)
pressure_min_entry.insert(0,0.9)
hyphen1 = tk.Label(frame, text="-",bg="lightgrey")
hyphen1.grid(row=1, column=2, padx=5)
pressure_max_entry = tk.Entry(frame, width=3)
pressure_max_entry.grid(row=1, column=3, padx=5)
pressure_max_entry.insert(0,1.2)

temperature_label = tk.Label(frame, text="Temperature",bg="lightgrey")
temperature_label.grid(row=1, column=4, padx=5, pady=5, sticky="e")
temp_min_entry = tk.Entry(frame, width=3)
temp_min_entry.grid(row=1, column=5, padx=5)
temp_min_entry.insert(0,5)
hyphen2 = tk.Label(frame, text="-",bg="lightgrey")
hyphen2.grid(row=1, column=6, padx=5)
temp_max_entry = tk.Entry(frame, width=3)
temp_max_entry.grid(row=1, column=7, padx=5)
temp_max_entry.insert(0,25)

bme_sensor_label = tk.Label(frame, text="BME Sensor - Pressure",bg="lightgrey")
bme_sensor_label.grid(row=2, column=0, padx=5, pady=5, sticky="e")
bme_pressure_min_entry = tk.Entry(frame, width=3)
bme_pressure_min_entry.grid(row=2, column=1, padx=5)
bme_pressure_min_entry.insert(0,0.9)
hyphen3 = tk.Label(frame, text="-",bg="lightgrey")
hyphen3.grid(row=2, column=2, padx=5)
bme_pressure_max_entry = tk.Entry(frame, width=3)
bme_pressure_max_entry.grid(row=2, column=3, padx=5)
bme_pressure_max_entry.insert(0,1.2)

bme_temperature_label = tk.Label(frame, text="Temperature",bg="lightgrey")
bme_temperature_label.grid(row=2, column=4, padx=5, pady=5, sticky="e")
bme_temp_min_entry = tk.Entry(frame, width=3)
bme_temp_min_entry.grid(row=2, column=5, padx=5)
bme_temp_min_entry.insert(0,5)
hyphen4 = tk.Label(frame, text="-",bg="lightgrey")
hyphen4.grid(row=2, column=6, padx=5)
bme_temp_max_entry = tk.Entry(frame, width=3)
bme_temp_max_entry.grid(row=2, column=7, padx=5)
bme_temp_max_entry.insert(0,25)

bme_altitude_label = tk.Label(frame, text="Altitude",bg="lightgrey")
bme_altitude_label.grid(row=2, column=8, padx=5, pady=5, sticky="e")
bme_altitude_min_entry = tk.Entry(frame, width=3)
bme_altitude_min_entry.grid(row=2, column=9, padx=5)
bme_altitude_min_entry.insert(0,200)
hyphen5 = tk.Label(frame, text="-",bg="lightgrey")
hyphen5.grid(row=2, column=10, padx=5)
bme_altitude_max_entry = tk.Entry(frame, width=3)
bme_altitude_max_entry.grid(row=2, column=11, padx=5)
bme_altitude_max_entry.insert(0,500)

battery_label = tk.Label(frame, text="Battery - Voltage",bg="lightgrey")
battery_label.grid(row=3, column=0, padx=5, pady=5, sticky="e")
battery_voltage_min_entry = tk.Entry(frame, width=3)
battery_voltage_min_entry.grid(row=3, column=1, padx=5)
battery_voltage_min_entry.insert(0,3)
hyphen6 = tk.Label(frame, text="-",bg="lightgrey")
hyphen6.grid(row=3, column=2, padx=5)
battery_voltage_max_entry = tk.Entry(frame, width=3)
battery_voltage_max_entry.grid(row=3, column=3, padx=5)
battery_voltage_max_entry.insert(0,3.5)

port_var2 = StringVar()
port_frame2 = tk.Frame(app)
port_frame2.pack(pady=5)
tk.Label(port_frame2, text="Select Debug Port - ",font=("Poppins", 12, "bold")).pack(side=tk.LEFT, padx=5)
port_combobox2 = ttk.Combobox(port_frame2, textvariable=port_var2, values=get_ports())
port_combobox2.pack(side=tk.LEFT)

connect_button2 = tk.Button(port_frame2, text="Connect", command=connect2)
connect_button2.pack(side=tk.LEFT, padx=5)

disconnect_button= tk.Button(port_frame2, text="Disconnect", command=disconnect2)
disconnect_button.pack(side=tk.LEFT, padx=5)

refresh_button2 = tk.Button(port_frame2, text="Refresh Ports", command=refresh_port2)
refresh_button2.pack(side=tk.LEFT, padx=5)

port_var1 = StringVar()
port_frame1 = tk.Frame(app)
port_frame1.pack(pady=5)
tk.Label(port_frame1, text="Select Debug Port - ",font=("Poppins", 12, "bold")).pack(side=tk.LEFT, padx=5)
port_combobox1 = ttk.Combobox(port_frame1, textvariable=port_var1, values=get_ports())
port_combobox1.pack(side=tk.LEFT)

connect_button1 = tk.Button(port_frame1, text="Connect", command=connect)
connect_button1.pack(side=tk.LEFT, padx=5)

refresh_button1 = tk.Button(port_frame1, text="Refresh Ports", command=refresh_port1)
refresh_button1.pack(side=tk.LEFT, padx=5)

frame0=tk.Frame(app)
pcbid_var = StringVar()
frame0.pack(pady=5)
pcbid_var.trace_add("write", enable_start_button)  
tk.Label(frame0, text="PCB ID - ",font=("Poppins", 12, "bold")).pack(side=tk.LEFT)
tk.Entry(frame0, textvariable=pcbid_var, width=20).pack(side=tk.LEFT, padx=5)

start_button = tk.Button(frame0, text="Start", command=start_test,state=tk.DISABLED)
start_button.pack(side=tk.LEFT, padx=5)

stop_button = tk.Button(frame0, text="Stop", command=stop_test,state=tk.DISABLED)
stop_button.pack(side=tk.LEFT, padx=5)

so_button = tk.Button(frame0, text="Start over",command=startover)
so_button.pack(side=tk.LEFT, padx=5)

frameS=tk.Frame(app)
frameS.pack(pady=5)

running_text=tk.Label(frameS,text="Running",font=("Poppins", 12, "bold"))
running_text.pack(side=tk.LEFT)
running_label = tk.Label(frameS, width=2, height=1, relief="solid", bg="white")
running_label.pack(side=tk.LEFT,padx=5)
qcdone_text=tk.Label(frameS,text="QC Done",font=("Poppins", 12, "bold"))
qcdone_text.pack(side=tk.LEFT)
qcdone_label = tk.Label(frameS, width=2, height=1, relief="solid", bg="white")
qcdone_label.pack(side=tk.LEFT)
stopped_text=tk.Label(frameS,text="Stopped",font=("Poppins", 12, "bold"))
stopped_text.pack(side=tk.LEFT)
stopped_label = tk.Label(frameS, width=2, height=1, relief="solid", bg="white")
stopped_label.pack(side=tk.LEFT)
time_label = tk.Label(frameS, text="Timer: 00:00", font=("Poppins", 12, "bold"))
time_label.pack(side=tk.LEFT,padx=5)

frame1 = tk.Frame(app,bd=3, relief="sunken",bg="lightgrey")
frame1.pack(fill="both",pady=5, padx=10)

text00 = tk.Label(frame1, text="Test",font=("Poppins", 12, "bold"),bg="grey",anchor="w")
text00.grid(row=0, column=0, sticky="nsew")
text01 = tk.Label(frame1, text="State",font=("Poppins", 12, "bold"),bg="grey")
text01.grid(row=0, column=1, sticky="nsew")
text02 = tk.Label(frame1, text="Accept/Reject",font=("Poppins", 12, "bold"),bg="grey")
text02.grid(row=0, column=2, sticky="nsew")
tk.Label(frame1, text="",font=("Poppins", 12, "bold"),bg="grey").grid(row=0, column=3, sticky="nsew")
tk.Label(frame1, text="",font=("Poppins", 12, "bold"),bg="grey").grid(row=0, column=4, sticky="nsew")

text10 = tk.Label(frame1, text="Watch Dog Timer",anchor="w",bg="lightgrey")
text10.grid(row=1, column=0,sticky="nsew")
text11 = tk.Label(frame1, text="--",bg="lightgrey")
text11.grid(row=1, column=1,sticky="nsew")
text12 = tk.Label(frame1, text="--",bg="lightgrey")
text12.grid(row=1, column=2,sticky="nsew")
b13=tk.Button(frame1,text="Accept",highlightbackground="lightgrey",command=wdt_accept)
b13.grid(row=1,column=3)
b14=tk.Button(frame1,text="Reject",highlightbackground="lightgrey",command=wdt_reject)
b14.grid(row=1,column=4)

text20 = tk.Label(frame1, text="ESP32 Flash Info",anchor="w",bg="lightgrey")
text20.grid(row=2, column=0,sticky="nsew")
text21 = tk.Label(frame1, text="--",bg="lightgrey")
text21.grid(row=2, column=1,sticky="nsew")
text22 = tk.Label(frame1, text="--",bg="lightgrey")
text22.grid(row=2, column=2,sticky="nsew")
b23=tk.Button(frame1,text="Accept",highlightbackground="lightgrey",command=info_accept)
b23.grid(row=2,column=3)
b24=tk.Button(frame1,text="Reject",highlightbackground="lightgrey",command=info_reject)
b24.grid(row=2,column=4)

text30 = tk.Label(frame1, text="RTC Communication",anchor="w",bg="lightgrey")
text30.grid(row=3, column=0,sticky="nsew")
text31 = tk.Label(frame1, text="--",bg="lightgrey")
text31.grid(row=3, column=1,sticky="nsew")
text32 = tk.Label(frame1, text="--",bg="lightgrey")
text32.grid(row=3, column=2,sticky="nsew")
b33=tk.Button(frame1,text="Accept",highlightbackground="lightgrey",command=rtc_accept)
b33.grid(row=3,column=3)
b34=tk.Button(frame1,text="Reject",highlightbackground="lightgrey",command=rtc_reject)
b34.grid(row=3,column=4)

text40 = tk.Label(frame1, text="Pressure Sensor Communication",anchor="w",bg="lightgrey")
text40.grid(row=4, column=0,sticky="nsew")
text41 = tk.Label(frame1, text="--",bg="lightgrey")
text41.grid(row=4, column=1,sticky="nsew")
text42 = tk.Label(frame1, text="--",bg="lightgrey")
text42.grid(row=4, column=2,sticky="nsew")
b43=tk.Button(frame1,text="Accept",highlightbackground="lightgrey",command=pressure_accept)
b43.grid(row=4,column=3)
b44=tk.Button(frame1,text="Reject",highlightbackground="lightgrey",command=pressure_reject)
b44.grid(row=4,column=4)

text50 = tk.Label(frame1, text="Battery Voltage",anchor="w",bg="lightgrey")
text50.grid(row=5, column=0,sticky="nsew")
text51 = tk.Label(frame1, text="--",bg="lightgrey")
text51.grid(row=5, column=1,sticky="nsew")
text52 = tk.Label(frame1, text="--",bg="lightgrey")
text52.grid(row=5, column=2,sticky="nsew")
b53=tk.Button(frame1,text="Accept",highlightbackground="lightgrey",command=battery_accept)
b53.grid(row=5,column=3)
b54=tk.Button(frame1,text="Reject",highlightbackground="lightgrey",command=battery_reject)
b54.grid(row=5,column=4)

text60 = tk.Label(frame1, text="SD Card Read+Write",anchor="w",bg="lightgrey")
text60.grid(row=6, column=0,sticky="nsew")
text61 = tk.Label(frame1, text="--",bg="lightgrey")
text61.grid(row=6, column=1,sticky="nsew")
text62 = tk.Label(frame1, text="--",bg="lightgrey")
text62.grid(row=6, column=2,sticky="nsew")
b63=tk.Button(frame1,text="Accept",highlightbackground="lightgrey",command=sd_accept)
b63.grid(row=6,column=3)
b64=tk.Button(frame1,text="Reject",highlightbackground="lightgrey",command=sd_reject)
b64.grid(row=6,column=4)

text70 = tk.Label(frame1, text="BME Sensor Communication",anchor="w",bg="lightgrey")
text70.grid(row=7, column=0,sticky="nsew")
text71 = tk.Label(frame1, text="--",bg="lightgrey")
text71.grid(row=7, column=1,sticky="nsew")
text72 = tk.Label(frame1, text="--",bg="lightgrey")
text72.grid(row=7, column=2,sticky="nsew")
b73=tk.Button(frame1,text="Accept",highlightbackground="lightgrey",command=bme_accept)
b73.grid(row=7,column=3)
b74=tk.Button(frame1,text="Reject",highlightbackground="lightgrey",command=bme_reject)
b74.grid(row=7,column=4)

text80 = tk.Label(frame1, text="GPRS Communication",anchor="w",bg="lightgrey")
text80.grid(row=8, column=0,sticky="nsew")
text81 = tk.Label(frame1, text="--",bg="lightgrey")
text81.grid(row=8, column=1,sticky="nsew")
text82 = tk.Label(frame1, text="--",bg="lightgrey")
text82.grid(row=8, column=2,sticky="nsew")
b83=tk.Button(frame1,text="Accept",highlightbackground="lightgrey",command=net_accept)
b83.grid(row=8,column=3)
b84=tk.Button(frame1,text="Reject",highlightbackground="lightgrey",command=net_reject)
b84.grid(row=8,column=4)

text90 = tk.Label(frame1, text="Hibernate Mode Testing",anchor="w",bg="lightgrey")
text90.grid(row=9, column=0,sticky="nsew")
text91 = tk.Label(frame1, text="--",bg="lightgrey")
text91.grid(row=9, column=1,sticky="nsew")
text92 = tk.Label(frame1, text="--",bg="lightgrey")
text92.grid(row=9, column=2,sticky="nsew")
b93=tk.Button(frame1,text="Accept",highlightbackground="lightgrey",command=hibernate_accept)
b93.grid(row=9,column=3)
b94=tk.Button(frame1,text="Reject",highlightbackground="lightgrey",command=hibernate_reject)
b94.grid(row=9,column=4)

frame1.grid_columnconfigure(0, weight=7)  
frame1.grid_columnconfigure(1, weight=1)  
frame1.grid_columnconfigure(2, weight=1)  
frame1.grid_columnconfigure(3, weight=1) 

remark_var = StringVar()
remark_frame = tk.Frame(app)
remark_frame.pack(pady=5)
tk.Label(remark_frame, text="Remark - ",font=("Poppins", 12, "bold")).pack(side=tk.LEFT)
tk.Entry(remark_frame, textvariable=remark_var, width=30).pack(side=tk.LEFT, padx=5)

label_frame = tk.Frame(app)
label_frame.pack(padx=10, pady=5)
pass_label = tk.Button(
label_frame, text="Pass",bg="white",command=passtest,height=3,width=20)
pass_label.pack(side=tk.LEFT,pady=5)
fail_label = tk.Button(
    label_frame, text="Fail",bg="white",command=failtest,height=3,width=20)
fail_label.pack(side=tk.LEFT,pady=5)

output_area = tk.Text(app, height=7, width=100, state=tk.NORMAL)
output_area.pack()

#----------------------------------------------------------------------------

values={"wdt":"--","info":"--","rtc":"--","pressure":"--","battery":"--","sd":"--","bme":"--","net":"--","hibernate":"--"}
ser = None #connect,start,stop,readfromserial,startover
is_port_running = False #enablestartbutton,connect,startover,readfromserial
hflag=None #readfromserial,startover
txt_filename=None #setuplogging,start
datee=None #stat,startover
completeflag=False #startover,readfromserial,stop
timee=None #start,startover
intime=True #readfromserial,starttimer
wdtflag=False #readfromserial
qcflag=None #readfromserial
trunning = False #updatetimer,starttest
counter = 0  #updatetimer,startover
mac=None #starttest
ser2=None #connect2,disconnect2,listen_for_command
data=None

refresh_port1()
refresh_port2()

app.mainloop()